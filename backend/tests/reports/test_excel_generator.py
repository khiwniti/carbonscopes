"""Tests for Excel report generator."""

import pytest
from pathlib import Path
from decimal import Decimal
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from reports.generators.excel_generator import ExcelReportGenerator


@pytest.fixture
def excel_generator():
    """Create Excel generator instance."""
    return ExcelReportGenerator()


@pytest.fixture
def sample_report_data():
    """Sample carbon analysis data for testing."""
    return {
        "analysis_id": "test-analysis-001",
        "project_name": "Test Construction Project",
        "boq_filename": "test_boq.xlsx",
        "tgo_version": "2026-03",
        "total_carbon": 125000.50,
        "material_count": 45,
        "matched_count": 42,
        "auto_matched_count": 38,
        "timestamp": datetime.now().isoformat(),
        "breakdown": [
            {
                "line_number": "01.01.001",
                "description_th": "คอนกรีตผสมเสร็จ fc 240 ksc",
                "description_en": "Ready-mix concrete fc 240 ksc",
                "quantity": 250.0,
                "unit": "ลบ.ม.",
                "carbon": 52500.0,
                "percentage": 42.0,
                "match_classification": "auto_matched"
            },
            {
                "line_number": "02.01.002",
                "description_th": "เหล็กเส้นกลม SR24",
                "description_en": "Round steel bar SR24",
                "quantity": 15000.0,
                "unit": "กก.",
                "carbon": 27000.0,
                "percentage": 21.6,
                "match_classification": "auto_matched"
            },
            {
                "line_number": "03.02.001",
                "description_th": "อิฐมอญ 4x8x16 ซม.",
                "description_en": "Clay brick 4x8x16 cm",
                "quantity": 50000.0,
                "unit": "ก้อน",
                "carbon": 12500.0,
                "percentage": 10.0,
                "match_classification": "auto_matched"
            },
        ],
        "statistics": {
            "auto_matched": 38,
            "review_required": 4,
            "rejected": 3,
            "auto_match_rate": 84.4
        }
    }


@pytest.fixture
def sample_audit_trail():
    """Sample audit trail data."""
    return {
        "metadata": {
            "analysis_id": "test-analysis-001",
            "calculator_version": "1.0.0",
            "tgo_version": "2026-03",
            "brightway_version": "2.4.6",
            "calculation_date": datetime.now().isoformat()
        },
        "material_calculations": [
            {
                "boq_line_number": "01.01.001",
                "description_th": "คอนกรีตผสมเสร็จ fc 240 ksc",
                "tgo_material_id": "tgo:concrete_240",
                "quantity": 250.0,
                "unit": "ลบ.ม.",
                "emission_factor_value": 210.0,
                "emission_factor_unit": "kgCO2e/m3",
                "carbon_result": 52500.0,
                "calculation_formula": "250.0 ลบ.ม. × 210.0 kgCO2e/m3 = 52500.0 kgCO2e",
                "match_confidence": 0.95
            },
            {
                "boq_line_number": "02.01.002",
                "description_th": "เหล็กเส้นกลม SR24",
                "tgo_material_id": "tgo:steel_sr24",
                "quantity": 15000.0,
                "unit": "กก.",
                "emission_factor_value": 1.8,
                "emission_factor_unit": "kgCO2e/kg",
                "carbon_result": 27000.0,
                "calculation_formula": "15000.0 กก. × 1.8 kgCO2e/kg = 27000.0 kgCO2e",
                "match_confidence": 0.92
            }
        ]
    }


class TestExcelGenerator:
    """Test suite for Excel report generator."""

    def test_generator_initialization(self, excel_generator):
        """Test Excel generator initializes correctly."""
        assert excel_generator is not None

    def test_generate_basic_report(self, excel_generator, sample_report_data, tmp_path):
        """Test generating basic Excel report without audit trail."""
        output_path = tmp_path / "carbon_report.xlsx"

        excel_bytes = excel_generator.generate_report(
            data=sample_report_data,
            output_path=output_path
        )

        # Verify Excel generated
        assert excel_bytes is not None
        assert len(excel_bytes) > 0
        assert output_path.exists()
        assert output_path.stat().st_size > 0

    def test_generate_report_with_audit_trail(
        self,
        excel_generator,
        sample_report_data,
        sample_audit_trail,
        tmp_path
    ):
        """Test generating Excel report with audit trail."""
        output_path = tmp_path / "carbon_report_audit.xlsx"

        excel_bytes = excel_generator.generate_report(
            data=sample_report_data,
            audit_trail=sample_audit_trail,
            output_path=output_path
        )

        assert excel_bytes is not None
        assert output_path.exists()

        # Load and verify workbook structure
        wb = load_workbook(output_path)
        sheet_names = wb.sheetnames

        assert "Summary" in sheet_names
        assert "Material Breakdown" in sheet_names
        assert "Audit Trail" in sheet_names
        assert "Methodology" in sheet_names

    def test_summary_sheet_content(self, excel_generator, sample_report_data, tmp_path):
        """Test Summary sheet has correct content."""
        output_path = tmp_path / "test_summary.xlsx"

        excel_generator.generate_report(
            data=sample_report_data,
            output_path=output_path
        )

        wb = load_workbook(output_path)
        ws = wb["Summary"]

        # Check title
        assert "Carbon Footprint" in ws['A1'].value

        # Check key metrics are present
        cell_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "Analysis ID:" in cell_values
        assert "Total Carbon Footprint (kgCO2e):" in cell_values
        assert "Top 10 Carbon Contributors" in cell_values

    def test_breakdown_sheet_content(self, excel_generator, sample_report_data, tmp_path):
        """Test Material Breakdown sheet has correct content."""
        output_path = tmp_path / "test_breakdown.xlsx"

        excel_generator.generate_report(
            data=sample_report_data,
            output_path=output_path
        )

        wb = load_workbook(output_path)
        ws = wb["Material Breakdown"]

        # Check headers
        assert ws['A3'].value == "Line #"
        assert ws['B3'].value == "Material Description (TH)"
        assert ws['F3'].value == "Carbon (kgCO2e)"

        # Check data rows (should have 3 materials from sample data)
        assert ws['A4'].value == "01.01.001"  # First material
        assert ws['B4'].value == "คอนกรีตผสมเสร็จ fc 240 ksc"

    def test_audit_trail_sheet_content(
        self,
        excel_generator,
        sample_report_data,
        sample_audit_trail,
        tmp_path
    ):
        """Test Audit Trail sheet has correct content."""
        output_path = tmp_path / "test_audit.xlsx"

        excel_generator.generate_report(
            data=sample_report_data,
            audit_trail=sample_audit_trail,
            output_path=output_path
        )

        wb = load_workbook(output_path)
        ws = wb["Audit Trail"]

        # Check title and metadata
        assert "Audit Trail" in ws['A1'].value
        assert "Analysis Metadata" in str(ws['A3'].value)

        # Check calculation headers
        cell_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "Material Calculations" in cell_values
        assert "Emission Factor" in cell_values
        assert "Formula" in cell_values

    def test_methodology_sheet_content(self, excel_generator, sample_report_data, tmp_path):
        """Test Methodology sheet has correct content."""
        output_path = tmp_path / "test_methodology.xlsx"

        excel_generator.generate_report(
            data=sample_report_data,
            output_path=output_path
        )

        wb = load_workbook(output_path)
        ws = wb["Methodology"]

        # Check section headings
        cell_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "Methodology and Assumptions" in cell_values
        assert "Calculation Methodology" in cell_values
        assert "Key Assumptions" in cell_values
        assert "Data Sources" in cell_values

    def test_number_formatting(self, excel_generator, sample_report_data, tmp_path):
        """Test that numbers are properly formatted."""
        output_path = tmp_path / "test_formatting.xlsx"

        excel_generator.generate_report(
            data=sample_report_data,
            output_path=output_path
        )

        wb = load_workbook(output_path)
        ws = wb["Material Breakdown"]

        # Check quantity cell formatting (should have number format)
        quantity_cell = ws['D4']  # First material quantity
        assert quantity_cell.number_format == '#,##0.00'

        # Check carbon cell formatting
        carbon_cell = ws['F4']  # First material carbon
        assert carbon_cell.number_format == '#,##0.00'

        # Check percentage formatting
        percentage_cell = ws['G4']  # First material percentage
        assert percentage_cell.number_format == '0.00%'

    def test_column_widths(self, excel_generator, sample_report_data, tmp_path):
        """Test that columns are auto-sized appropriately."""
        output_path = tmp_path / "test_widths.xlsx"

        excel_generator.generate_report(
            data=sample_report_data,
            output_path=output_path
        )

        wb = load_workbook(output_path)
        ws = wb["Material Breakdown"]

        # Check that columns have been sized (not default width)
        for column in ws.column_dimensions.values():
            if column.width:
                assert column.width > 8  # Default width is 8.43

    def test_cell_styling(self, excel_generator, sample_report_data, tmp_path):
        """Test that cells have proper styling."""
        output_path = tmp_path / "test_styling.xlsx"

        excel_generator.generate_report(
            data=sample_report_data,
            output_path=output_path
        )

        wb = load_workbook(output_path)
        ws = wb["Summary"]

        # Check header cell has styling
        header_cell = ws['A1']
        assert header_cell.font.bold is True
        assert header_cell.font.size >= 14

    def test_generate_without_output_path(self, excel_generator, sample_report_data):
        """Test generating Excel without saving to file."""
        excel_bytes = excel_generator.generate_report(data=sample_report_data)

        # Should still return Excel bytes
        assert excel_bytes is not None
        assert len(excel_bytes) > 0

        # Verify it's a valid Excel file
        wb = load_workbook(BytesIO(excel_bytes))
        assert len(wb.sheetnames) >= 3

    def test_top_contributors_calculation(self, excel_generator, sample_report_data, tmp_path):
        """Test that top contributors are correctly sorted and limited to 10."""
        output_path = tmp_path / "test_top.xlsx"

        excel_generator.generate_report(
            data=sample_report_data,
            output_path=output_path
        )

        wb = load_workbook(output_path)
        ws = wb["Summary"]

        # Find the top contributors section and count rows
        contributors_count = 0
        in_contributors = False

        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "Top 10 Carbon Contributors":
                    in_contributors = True
                    break

            if in_contributors and row[0].value and isinstance(row[0].value, int):
                contributors_count += 1

        # Should have 3 contributors from sample data
        assert contributors_count == 3

    def test_total_row_in_breakdown(self, excel_generator, sample_report_data, tmp_path):
        """Test that breakdown sheet has a TOTAL row."""
        output_path = tmp_path / "test_total.xlsx"

        excel_generator.generate_report(
            data=sample_report_data,
            output_path=output_path
        )

        wb = load_workbook(output_path)
        ws = wb["Material Breakdown"]

        # Find TOTAL row
        total_found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "TOTAL:":
                    total_found = True
                    # Check that the carbon total is correct
                    carbon_cell = ws.cell(row=cell.row, column=6)
                    assert carbon_cell.value == 125000.50
                    break

        assert total_found

    def test_thai_unicode_support(self, excel_generator, sample_report_data, tmp_path):
        """Test that Thai characters are properly saved."""
        output_path = tmp_path / "test_thai.xlsx"

        excel_generator.generate_report(
            data=sample_report_data,
            output_path=output_path
        )

        wb = load_workbook(output_path)
        ws = wb["Material Breakdown"]

        # Check Thai description is preserved
        thai_text = ws['B4'].value
        assert thai_text == "คอนกรีตผสมเสร็จ fc 240 ksc"
        # Verify Thai Unicode range
        assert any(ord(c) >= 0x0E00 and ord(c) <= 0x0E7F for c in thai_text)

    def test_performance_basic_report(self, excel_generator, sample_report_data, benchmark):
        """Benchmark basic Excel report generation performance."""
        result = benchmark(
            excel_generator.generate_report,
            data=sample_report_data
        )

        assert result is not None
        assert len(result) > 0

    @pytest.mark.slow
    def test_large_dataset_performance(self, excel_generator, tmp_path):
        """Test Excel generation with large dataset (200+ materials)."""
        # Create large dataset
        large_data = {
            "analysis_id": "large-test",
            "project_name": "Large Project",
            "total_carbon": 2000000.0,
            "material_count": 200,
            "matched_count": 195,
            "timestamp": datetime.now().isoformat(),
            "breakdown": [
                {
                    "line_number": f"{i:06d}",
                    "description_th": f"วัสดุที่ {i}",
                    "description_en": f"Material {i}",
                    "quantity": 100.0 * i,
                    "unit": "หน่วย",
                    "carbon": 500.0 * i,
                    "percentage": 0.5,
                    "match_classification": "auto_matched"
                }
                for i in range(1, 201)
            ],
            "statistics": {"auto_matched": 195}
        }

        output_path = tmp_path / "large_report.xlsx"

        excel_bytes = excel_generator.generate_report(
            data=large_data,
            output_path=output_path
        )

        assert excel_bytes is not None
        assert output_path.exists()

        # Verify all 200 materials are in breakdown
        wb = load_workbook(output_path)
        ws = wb["Material Breakdown"]

        # Count data rows (excluding header and total)
        data_rows = sum(1 for row in ws.iter_rows(min_row=4) if row[0].value and row[0].value != "TOTAL:")
        assert data_rows == 200

    def test_empty_breakdown(self, excel_generator, tmp_path):
        """Test handling of empty breakdown list."""
        empty_data = {
            "analysis_id": "empty-test",
            "project_name": "Empty Project",
            "total_carbon": 0.0,
            "material_count": 0,
            "matched_count": 0,
            "timestamp": datetime.now().isoformat(),
            "breakdown": [],
            "statistics": {"auto_matched": 0}
        }

        output_path = tmp_path / "empty_report.xlsx"

        excel_bytes = excel_generator.generate_report(
            data=empty_data,
            output_path=output_path
        )

        assert excel_bytes is not None
        assert output_path.exists()

        wb = load_workbook(output_path)
        assert "Summary" in wb.sheetnames
        assert "Material Breakdown" in wb.sheetnames
