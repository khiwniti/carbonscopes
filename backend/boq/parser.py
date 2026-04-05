"""
Thai BOQ Excel Parser.

Handles complex Thai construction BOQ files with merged cells, bilingual headers,
and mixed unit formats.
"""

import openpyxl
import hashlib
import logging
from pathlib import Path
from decimal import Decimal, InvalidOperation
from typing import Optional
from datetime import datetime

from .models import BOQMaterial, BOQParseResult
from .unit_normalizer import normalize_unit

logger = logging.getLogger(__name__)


def detect_header_row(sheet) -> Optional[int]:
    """
    Detect BOQ header row by searching for Thai/English keywords.

    Searches for: รายการ, หน่วย, จำนวน (Thai) or
                  description, unit, quantity (English)

    Returns row index (0-based) or None if not found.
    """
    thai_keywords = ['รายการ', 'หน่วย', 'จำนวน']
    english_keywords = ['description', 'unit', 'quantity']

    for row_idx, row in enumerate(sheet.iter_rows(max_row=20)):
        row_text = ' '.join([
            str(cell.value).lower() if cell.value else ''
            for cell in row
        ])

        # Check for Thai keywords
        if any(kw in row_text for kw in thai_keywords):
            logger.info(f"Header row detected at line {row_idx + 1} (Thai)")
            return row_idx

        # Check for English keywords
        if any(kw in row_text for kw in english_keywords):
            logger.info(f"Header row detected at line {row_idx + 1} (English)")
            return row_idx

    return None


def propagate_merged_cells(sheet) -> dict:
    """
    Create mapping of merged cell values.
    Returns dict: {(row, col): cell_value}
    """
    merged_map = {}

    for merged_range in sheet.merged_cells.ranges:
        # Get top-left cell value
        top_row = merged_range.min_row
        top_col = merged_range.min_col
        value = sheet.cell(top_row, top_col).value

        # Map all cells in merged range to this value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_map[(row, col)] = value

    return merged_map


def parse_boq(
    file_path: str,
    encoding: str = 'utf-8'
) -> BOQParseResult:
    """
    Parse Thai BOQ Excel file.

    Args:
        file_path: Absolute path to .xlsx or .xls file
        encoding: Character encoding (default utf-8)

    Returns:
        BOQParseResult with parsed materials and metadata

    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file format is invalid
    """
    # Generate file hash (for caching key)
    with open(file_path, 'rb') as f:
        file_content = f.read()
        file_id = hashlib.sha256(file_content).hexdigest()

    filename = Path(file_path).name

    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # Detect header row
        header_row_idx = detect_header_row(sheet)
        if header_row_idx is None:
            raise ValueError("Could not detect header row (no Thai/English keywords found)")

        # Propagate merged cell values
        merged_map = propagate_merged_cells(sheet)

        # Parse materials
        materials = []
        errors = []
        total_lines = 0

        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 2)):
            total_lines += 1
            line_number = row_idx + header_row_idx + 2

            try:
                # Extract material data
                # Assume columns: [Item No, Description (Thai), Description (English), Unit, Quantity]
                # Adjust column indices based on actual BOQ format

                description_th = row[1].value if len(row) > 1 else None
                description_en = row[2].value if len(row) > 2 else None
                unit_raw = row[3].value if len(row) > 3 else None
                quantity_raw = row[4].value if len(row) > 4 else None

                # Skip empty rows
                if not description_th and not description_en:
                    continue

                # Normalize unit
                if not unit_raw:
                    errors.append({
                        "error_type": "UNIT_ERROR",
                        "line_number": line_number,
                        "column": "D",
                        "message_en": "Missing unit",
                        "message_th": "ไม่พบหน่วย",
                        "suggestion": "Add unit (e.g., ลบ.ม., ตร.ม., กก.)",
                        "timestamp": datetime.now().isoformat()
                    })
                    continue

                try:
                    unit, conversion_factor = normalize_unit(str(unit_raw))
                except ValueError:
                    errors.append({
                        "error_type": "UNIT_ERROR",
                        "line_number": line_number,
                        "column": "D",
                        "message_en": f"Unrecognized unit: {unit_raw}",
                        "message_th": f"ไม่รู้จักหน่วย: {unit_raw}",
                        "suggestion": "Use standard unit (ลบ.ม., ตร.ม., กก., ตัน)",
                        "timestamp": datetime.now().isoformat()
                    })
                    continue

                # Parse quantity
                try:
                    quantity = Decimal(str(quantity_raw))
                except (InvalidOperation, ValueError, TypeError):
                    errors.append({
                        "error_type": "PARSE_ERROR",
                        "line_number": line_number,
                        "column": "E",
                        "message_en": f"Invalid quantity: {quantity_raw}",
                        "message_th": f"จำนวนไม่ถูกต้อง: {quantity_raw}",
                        "suggestion": "Enter numeric value",
                        "timestamp": datetime.now().isoformat()
                    })
                    continue

                # Create material entry
                material = BOQMaterial(
                    line_number=line_number,
                    description_th=str(description_th),
                    description_en=str(description_en) if description_en else None,
                    quantity=quantity,
                    unit=unit,
                    unit_raw=str(unit_raw),
                    conversion_factor=conversion_factor
                )

                materials.append(material)

            except Exception as e:
                logger.warning(f"Error parsing line {line_number}: {e}")
                errors.append({
                    "error_type": "PARSE_ERROR",
                    "line_number": line_number,
                    "message_en": f"Parse error: {str(e)}",
                    "message_th": f"ข้อผิดพลาดในการแปลง: {str(e)}",
                    "timestamp": datetime.now().isoformat()
                })

        # Calculate success rate
        parsed_count = len(materials)
        success_rate = (parsed_count / total_lines * 100) if total_lines > 0 else 0

        # Determine status
        if parsed_count == 0:
            status = "failed"
        elif len(errors) > 0:
            status = "partial"
        else:
            status = "parsed"

        return BOQParseResult(
            file_id=file_id,
            filename=filename,
            status=status,
            materials=materials,
            metadata={
                "total_lines": total_lines,
                "parsed_lines": parsed_count,
                "success_rate": round(success_rate, 2),
                "error_count": len(errors)
            },
            errors=errors
        )

    except Exception as e:
        logger.error(f"Failed to parse BOQ file {filename}: {e}")
        raise ValueError(f"BOQ parsing failed: {str(e)}")
