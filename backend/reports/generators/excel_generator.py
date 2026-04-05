"""Excel Report Generator using openpyxl.

Generates professional Excel workbooks with:
- Multiple worksheets (Summary, Details, Audit Trail)
- Formulas and data validation
- Conditional formatting
- Professional styling
"""

import logging
from pathlib import Path
from typing import Dict, Any, Optional, List
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment,
    NamedStyle, Color
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import BarChart, PieChart, Reference

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generate Excel reports for carbon analysis results."""

    # Style definitions
    HEADER_STYLE = {
        'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

    TITLE_STYLE = {
        'font': Font(name='Calibri', size=14, bold=True, color='366092'),
        'alignment': Alignment(horizontal='left', vertical='center'),
    }

    DATA_STYLE = {
        'font': Font(name='Calibri', size=10),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
    }

    NUMBER_STYLE = {
        'font': Font(name='Calibri', size=10),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'border': Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
    }

    def __init__(self):
        """Initialize Excel generator."""
        logger.info("ExcelReportGenerator initialized")

    def generate_report(
        self,
        data: Dict[str, Any],
        audit_trail: Optional[Dict[str, Any]] = None,
        output_path: Optional[Path] = None
    ) -> bytes:
        """Generate complete Excel report with multiple sheets.

        Args:
            data: Report data containing:
                - analysis_id: UUID
                - project_name: Project name
                - total_carbon: Total kgCO2e
                - breakdown: List of material breakdowns
                - statistics: Calculation statistics
            audit_trail: Complete audit trail from calculation pipeline
            output_path: Optional path to save Excel file

        Returns:
            Excel file bytes
        """
        logger.info(f"Generating Excel report for analysis: {data.get('analysis_id')}")

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Add worksheets
        self._create_summary_sheet(wb, data)
        self._create_breakdown_sheet(wb, data)

        if audit_trail:
            self._create_audit_trail_sheet(wb, audit_trail)

        self._create_methodology_sheet(wb)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        excel_bytes = output.getvalue()

        # Save to file if path provided
        if output_path:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_bytes(excel_bytes)
            logger.info(f"Excel report saved to: {output_path}")

        return excel_bytes

    def _create_summary_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create summary worksheet with key metrics.

        Args:
            wb: Workbook instance
            data: Report data
        """
        ws = wb.create_sheet("Summary")

        # Title
        ws['A1'] = "Carbon Footprint Analysis - Summary"
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='366092')
        ws.merge_cells('A1:D1')

        # Metadata
        row = 3
        metadata = [
            ("Analysis ID:", data.get("analysis_id", "N/A")),
            ("Project Name:", data.get("project_name", "Unnamed Project")),
            ("BOQ Filename:", data.get("boq_filename", "N/A")),
            ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("TGO Version:", data.get("tgo_version", "2026-03")),
        ]

        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        # Key Results
        row += 2
        ws[f'A{row}'] = "Key Results"
        self._apply_style(ws[f'A{row}'], self.TITLE_STYLE)
        row += 1

        # Results table
        total_carbon = float(data.get("total_carbon", 0))
        material_count = data.get("material_count", 0)
        matched_count = data.get("matched_count", 0)
        auto_matched = data.get("statistics", {}).get("auto_matched", 0)

        results = [
            ("Total Carbon Footprint (kgCO2e):", total_carbon),
            ("Total Carbon Footprint (tCO2e):", total_carbon / 1000),
            ("Total Materials:", material_count),
            ("Matched Materials:", matched_count),
            ("Auto-Matched:", auto_matched),
            ("Match Rate (%):", (matched_count / material_count * 100) if material_count > 0 else 0),
        ]

        for label, value in results:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            if isinstance(value, (int, float)):
                ws[f'B{row}'].number_format = '#,##0.00'
            row += 1

        # Top Contributors
        row += 2
        ws[f'A{row}'] = "Top 10 Carbon Contributors"
        self._apply_style(ws[f'A{row}'], self.TITLE_STYLE)
        row += 1

        # Headers
        headers = ["Rank", "Material Description", "Quantity", "Unit", "Carbon (kgCO2e)", "% of Total"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.HEADER_STYLE)
        row += 1

        # Sort breakdown by carbon (descending)
        breakdown = data.get("breakdown", [])
        sorted_breakdown = sorted(
            [item for item in breakdown if item.get("carbon")],
            key=lambda x: float(x["carbon"]) if x.get("carbon") else 0,
            reverse=True
        )[:10]

        # Data rows
        for rank, item in enumerate(sorted_breakdown, start=1):
            carbon = float(item.get("carbon", 0))
            percentage = (carbon / total_carbon * 100) if total_carbon > 0 else 0

            ws.cell(row=row, column=1, value=rank)
            ws.cell(row=row, column=2, value=item.get("description_th", "N/A"))
            ws.cell(row=row, column=3, value=float(item.get("quantity", 0)))
            ws.cell(row=row, column=4, value=item.get("unit", ""))
            ws.cell(row=row, column=5, value=carbon)
            ws.cell(row=row, column=6, value=percentage)

            # Apply number formatting
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6).number_format = '0.00%'

            # Apply styles
            for col in range(1, 7):
                if col in [3, 5, 6]:
                    self._apply_style(ws.cell(row=row, column=col), self.NUMBER_STYLE)
                else:
                    self._apply_style(ws.cell(row=row, column=col), self.DATA_STYLE)

            row += 1

        # Auto-fit columns
        self._auto_fit_columns(ws)

    def _create_breakdown_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create detailed material breakdown worksheet.

        Args:
            wb: Workbook instance
            data: Report data
        """
        ws = wb.create_sheet("Material Breakdown")

        # Title
        ws['A1'] = "Complete Material Breakdown"
        self._apply_style(ws['A1'], self.TITLE_STYLE)
        ws.merge_cells('A1:H1')

        # Headers
        row = 3
        headers = [
            "Line #",
            "Material Description (TH)",
            "Material Description (EN)",
            "Quantity",
            "Unit",
            "Carbon (kgCO2e)",
            "% of Total",
            "Match Status"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.HEADER_STYLE)
        row += 1

        # Data rows
        breakdown = data.get("breakdown", [])
        total_carbon = float(data.get("total_carbon", 0))

        for item in breakdown:
            carbon = float(item.get("carbon", 0)) if item.get("carbon") else 0
            percentage = (carbon / total_carbon * 100) if total_carbon > 0 else 0

            ws.cell(row=row, column=1, value=item.get("line_number", ""))
            ws.cell(row=row, column=2, value=item.get("description_th", ""))
            ws.cell(row=row, column=3, value=item.get("description_en", ""))
            ws.cell(row=row, column=4, value=float(item.get("quantity", 0)))
            ws.cell(row=row, column=5, value=item.get("unit", ""))
            ws.cell(row=row, column=6, value=carbon)
            ws.cell(row=row, column=7, value=percentage / 100)  # Excel percentage
            ws.cell(row=row, column=8, value=item.get("match_classification", "N/A"))

            # Number formatting
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            ws.cell(row=row, column=7).number_format = '0.00%'

            # Apply styles
            for col in range(1, 9):
                if col in [4, 6, 7]:
                    self._apply_style(ws.cell(row=row, column=col), self.NUMBER_STYLE)
                else:
                    self._apply_style(ws.cell(row=row, column=col), self.DATA_STYLE)

            row += 1

        # Total row
        ws.cell(row=row, column=5, value="TOTAL:")
        ws.cell(row=row, column=5).font = Font(bold=True)
        ws.cell(row=row, column=6, value=total_carbon)
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        ws.cell(row=row, column=6).font = Font(bold=True)
        ws.cell(row=row, column=7, value=1.0)
        ws.cell(row=row, column=7).number_format = '0.00%'
        ws.cell(row=row, column=7).font = Font(bold=True)

        # Auto-fit columns
        self._auto_fit_columns(ws)

    def _create_audit_trail_sheet(self, wb: Workbook, audit_trail: Dict[str, Any]) -> None:
        """Create audit trail worksheet with calculation details.

        Args:
            wb: Workbook instance
            audit_trail: Complete audit trail data
        """
        ws = wb.create_sheet("Audit Trail")

        # Title
        ws['A1'] = "Calculation Audit Trail"
        self._apply_style(ws['A1'], self.TITLE_STYLE)
        ws.merge_cells('A1:J1')

        # Metadata
        row = 3
        ws[f'A{row}'] = "Analysis Metadata"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        metadata_items = audit_trail.get("metadata", {})
        for key, value in metadata_items.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value)
            row += 1

        # Material calculations
        row += 2
        ws[f'A{row}'] = "Material Calculations"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        # Headers
        headers = [
            "Line #",
            "Material",
            "TGO Material ID",
            "Quantity",
            "Unit",
            "Emission Factor",
            "EF Unit",
            "Carbon Result",
            "Formula",
            "Confidence"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.HEADER_STYLE)
        row += 1

        # Material calculation rows
        material_calcs = audit_trail.get("material_calculations", [])
        for calc in material_calcs:
            ws.cell(row=row, column=1, value=calc.get("boq_line_number", ""))
            ws.cell(row=row, column=2, value=calc.get("description_th", ""))
            ws.cell(row=row, column=3, value=calc.get("tgo_material_id", ""))
            ws.cell(row=row, column=4, value=float(calc.get("quantity", 0)))
            ws.cell(row=row, column=5, value=calc.get("unit", ""))
            ws.cell(row=row, column=6, value=float(calc.get("emission_factor_value", 0)) if calc.get("emission_factor_value") else "")
            ws.cell(row=row, column=7, value=calc.get("emission_factor_unit", ""))
            ws.cell(row=row, column=8, value=float(calc.get("carbon_result", 0)) if calc.get("carbon_result") else "")
            ws.cell(row=row, column=9, value=calc.get("calculation_formula", ""))
            ws.cell(row=row, column=10, value=float(calc.get("match_confidence", 0)) if calc.get("match_confidence") else "")

            # Number formatting
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=6).number_format = '#,##0.0000'
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            ws.cell(row=row, column=10).number_format = '0.00%'

            row += 1

        # Auto-fit columns
        self._auto_fit_columns(ws)

    def _create_methodology_sheet(self, wb: Workbook) -> None:
        """Create methodology and assumptions worksheet.

        Args:
            wb: Workbook instance
        """
        ws = wb.create_sheet("Methodology")

        # Title
        ws['A1'] = "Methodology and Assumptions"
        self._apply_style(ws['A1'], self.TITLE_STYLE)
        ws.merge_cells('A1:E1')

        # Methodology
        row = 3
        ws[f'A{row}'] = "Calculation Methodology"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2

        methodology_text = """
This carbon footprint analysis follows Life Cycle Assessment (LCA) methodology
according to ISO 14040 and ISO 14044 standards.

Carbon calculations use emission factors from the TGO (Thailand Greenhouse Gas
Management Organization) database and Brightway2 LCA framework.

The analysis covers embodied carbon (cradle-to-gate) for construction materials
as specified in the Bill of Quantities (BOQ).
        """

        ws[f'A{row}'] = methodology_text
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:E{row+5}')
        ws.row_dimensions[row].height = 120

        # Key Assumptions
        row += 7
        ws[f'A{row}'] = "Key Assumptions"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        assumptions = [
            "1. Emission factors from TGO database (2026 version)",
            "2. Embodied carbon only (cradle-to-gate) for construction materials",
            "3. Transportation and installation emissions not included",
            "4. Average values used for materials without specific brand/source",
            "5. Automated material matching uses fuzzy matching with confidence ≥ 0.70",
            "6. Materials with match confidence < 0.70 require manual review",
            "7. Calculations follow GWP100 (Global Warming Potential over 100 years)",
            "8. System boundary: Raw material extraction to factory gate",
        ]

        for assumption in assumptions:
            ws[f'A{row}'] = assumption
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1

        # Data Sources
        row += 2
        ws[f'A{row}'] = "Data Sources"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        sources = [
            ("TGO Database:", "Thailand Greenhouse Gas Management Organization"),
            ("LCA Framework:", "Brightway2 (open-source)"),
            ("Standards:", "ISO 14040, ISO 14044"),
            ("BOQ Format:", "Thai construction industry standard"),
        ]

        for label, value in sources:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        # Auto-fit columns
        self._auto_fit_columns(ws)

    def _apply_style(self, cell, style_dict: Dict[str, Any]) -> None:
        """Apply style dictionary to cell.

        Args:
            cell: Cell object
            style_dict: Style parameters
        """
        for key, value in style_dict.items():
            setattr(cell, key, value)

    def _auto_fit_columns(self, ws: Worksheet) -> None:
        """Auto-fit column widths based on content.

        Args:
            ws: Worksheet instance
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Set column width (with max cap)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
